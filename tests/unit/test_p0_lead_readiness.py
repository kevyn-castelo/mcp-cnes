from __future__ import annotations

import csv
import json
from pathlib import Path

import pytest
from mcp import Client
from openpyxl import load_workbook

from mcp_cnes.domain.models import HospitalInfo
from mcp_cnes.infrastructure.config import Settings
from mcp_cnes.infrastructure.persistence import MemoryCNESRepository
from mcp_cnes.interfaces.mcp import create_mcp_server


def lead_hospital(index: int) -> HospitalInfo:
    return HospitalInfo(
        cnes=f"{index:07d}",
        nome_fantasia=f"Hospital {index}",
        municipio="São Paulo",
        uf="SP",
        tipo_estabelecimento="05 - HOSPITAL GERAL",
        natureza_juridica="2062 - Sociedade Empresária",
        gestao="M",
        convenio_sus=index % 2 == 0,
        leitos_existentes=index * 10,
        leitos_sus=index * 5 if index % 2 == 0 else 0,
        competencia="202501",
    )


def repository_with_leads() -> MemoryCNESRepository:
    repository = MemoryCNESRepository()
    repository.replace_all(
        [lead_hospital(index) for index in range(1, 13)],
        "leads.csv",
        batch_id="lead-batch",
    )
    repository.update_batch_metadata(
        "lead-batch",
        "portal_sus_hospitais_leitos",
        "202501",
        {"municipio": "São Paulo"},
        '"etag-fixture"',
    )
    return repository


@pytest.mark.asyncio
async def test_simple_searches_share_filters_and_default_to_bed_size_descending(
    tmp_path: Path,
) -> None:
    server = create_mcp_server(
        settings=Settings(database_path=tmp_path / "unused.sqlite3"),
        repository=repository_with_leads(),
    )

    async with Client(server) as client:
        municipality = await client.call_tool(
            "cnes_search_municipio",
            {
                "municipio": "São Paulo",
                "tipo_estabelecimento": "HOSPITAL",
                "order_by": "leitos_existentes",
                "limit": 10,
            },
        )
        state = await client.call_tool(
            "cnes_search_uf",
            {
                "uf": "SP",
                "municipio": "Paulo",
                "tipo_estabelecimento": "HOSPITAL",
                "natureza_juridica": "Empresária",
                "gestao": "M",
                "convenio_sus": True,
                "limit": 10,
            },
        )

    municipality_beds = [
        item["leitos_existentes"] for item in municipality.structured_content["estabelecimentos"]
    ]
    state_beds = [
        item["leitos_existentes"] for item in state.structured_content["estabelecimentos"]
    ]
    assert municipality_beds == [120, 110, 100, 90, 80, 70, 60, 50, 40, 30]
    assert state_beds == [120, 100, 80, 60, 40, 20]


@pytest.mark.asyncio
async def test_export_explicit_cnes_slice_has_exact_rows_and_provenance(
    tmp_path: Path,
) -> None:
    server = create_mcp_server(
        settings=Settings(
            database_path=tmp_path / "unused.sqlite3",
            output_dir=tmp_path / "exports",
        ),
        repository=repository_with_leads(),
    )
    selected = [f"{index:07d}" for index in range(1, 11)]

    async with Client(server) as client:
        csv_result = await client.call_tool(
            "cnes_export",
            {"formato": "csv", "cnes_list": selected},
        )
        xlsx_result = await client.call_tool(
            "cnes_export",
            {
                "formato": "xlsx",
                "cnes_list": selected,
                "limit": 10,
                "offset": 0,
            },
        )
        sliced_result = await client.call_tool(
            "cnes_export",
            {"formato": "json", "limit": 3, "offset": 2, "order_by": "cnes"},
        )

    assert csv_result.structured_content["registros"] == 10
    with Path(csv_result.structured_content["filepath"]).open(
        encoding="utf-8", newline=""
    ) as handle:
        rows = list(csv.DictReader(handle))
    assert len(rows) == 10
    assert {row["CNES"] for row in rows} == set(selected)
    metadata = rows[0]
    assert metadata["_competencia"] == "202501"
    assert metadata["_lote_id"] == "lead-batch"
    assert metadata["_etag"] == '"etag-fixture"'
    assert metadata["_versao_contrato"] == "v1"
    assert json.loads(metadata["_filtros_aplicados"])["cnes_list"] == selected
    assert metadata["_extraido_em"]

    sliced = json.loads(
        Path(sliced_result.structured_content["filepath"]).read_text(encoding="utf-8")
    )
    assert [item["cnes"] for item in sliced] == ["0000003", "0000004", "0000005"]

    workbook = load_workbook(Path(xlsx_result.structured_content["filepath"]), read_only=True)
    try:
        assert workbook.sheetnames == ["CNES", "_metadados"]
        values = {
            row[0]: row[1] for row in workbook["_metadados"].iter_rows(min_row=2, values_only=True)
        }
        assert values["lote_id"] == "lead-batch"
        assert values["etag"] == '"etag-fixture"'
        assert values["versao_contrato"] == "v1"
    finally:
        workbook.close()


@pytest.mark.asyncio
async def test_export_freezes_active_batch_before_streaming_pages(tmp_path: Path) -> None:
    class SwitchingRepository(MemoryCNESRepository):
        def __init__(self) -> None:
            super().__init__()
            self.batch_ids_used: list[str | None] = []

        def get_batch_metadata(self, batch_id: str | None = None):
            metadata = super().get_batch_metadata(batch_id)
            self.activate_batch("other-batch")
            return metadata

        def advanced_search(
            self,
            filters,
            order_by,
            offset,
            limit,
            batch_id=None,
        ):
            self.batch_ids_used.append(batch_id)
            return super().advanced_search(filters, order_by, offset, limit, batch_id)

    repository = SwitchingRepository()
    repository.replace_all(
        [lead_hospital(index) for index in range(1, 13)],
        "leads.csv",
        batch_id="lead-batch",
    )
    repository.update_batch_metadata("lead-batch", "portal_sus_hospitais_leitos", "202501", {})
    repository.replace_all([lead_hospital(99)], "other.csv", batch_id="other-batch")
    repository.activate_batch("lead-batch")
    server = create_mcp_server(
        settings=Settings(
            database_path=tmp_path / "unused.sqlite3",
            output_dir=tmp_path / "exports",
        ),
        repository=repository,
    )

    async with Client(server) as client:
        result = await client.call_tool("cnes_export", {"formato": "json", "limit": 2})

    rows = json.loads(Path(result.structured_content["filepath"]).read_text(encoding="utf-8"))
    assert [row["cnes"] for row in rows] == ["0000001", "0000002"]
    assert rows[0]["_metadados"]["lote_id"] == "lead-batch"
    assert repository.batch_ids_used == ["lead-batch"]
