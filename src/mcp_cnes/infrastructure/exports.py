"""Exportação atômica de datasets canônicos para arquivos locais."""

from __future__ import annotations

import csv
import json
import os
import tempfile
from collections.abc import Iterable, Mapping
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from mcp_cnes.domain.models import HospitalInfo

EXPORT_COLUMNS = {
    "COMPETENCIA": "competencia",
    "UF": "uf",
    "MUNICIPIO": "municipio",
    "CNES": "cnes",
    "NOME_FANTASIA": "nome_fantasia",
    "TIPO_ESTABELECIMENTO": "tipo_estabelecimento",
    "NATUREZA_JURIDICA": "natureza_juridica",
    "GESTAO": "gestao",
    "CONVENIO_SUS": "convenio_sus",
    "LEITOS_EXISTENTES": "leitos_existentes",
    "LEITOS_SUS": "leitos_sus",
}
CRM_COLUMNS = (
    "chave_deduplicacao",
    "cnes",
    "cnpj",
    "cnpj_mantenedora",
    "razao_social",
    "nome_fantasia",
    "tipo_estabelecimento",
    "natureza_juridica",
    "gestao",
    "municipio",
    "uf",
    "logradouro",
    "numero",
    "complemento",
    "bairro",
    "cep",
    "telefone",
    "email",
    "leitos_existentes",
    "leitos_sus",
    "leitos_uti_adulto",
    "leitos_uti_pediatrica",
    "leitos_uti_neonatal",
    "leitos_cirurgicos",
    "leitos_clinicos",
    "leitos_obstetricos",
    "leitos_complementares",
    "habilitacoes",
    "total_habilitacoes",
    "campos_ausentes",
    "competencia",
)


class LocalDatasetExporter:
    def __init__(self, output_dir: Path) -> None:
        self._root = output_dir.resolve(strict=False)

    def export(
        self,
        hospitals: Iterable[HospitalInfo],
        format: str,
        destination: Path | None,
        basename: str,
        metadata: Mapping[str, Any] | None = None,
        output_profile: str | None = None,
    ) -> tuple[Path, int]:
        normalized_format = format.casefold()
        if normalized_format not in {"csv", "json", "jsonl", "xlsx"}:
            raise ValueError("formato deve ser csv, json, jsonl ou xlsx")
        if output_profile not in {None, "crm_generico"}:
            raise ValueError("perfil de saída não suportado")
        directory = self._resolve_destination(destination)
        directory.mkdir(parents=True, exist_ok=True)
        output = self._available_output(directory, basename, normalized_format)
        temporary: Path | None = None
        records = 0
        try:
            descriptor, name = tempfile.mkstemp(
                prefix=f".{basename}.", suffix=f".{normalized_format}", dir=directory
            )
            os.close(descriptor)
            temporary = Path(name)
            if normalized_format == "csv":
                with temporary.open("w", encoding="utf-8", newline="") as handle:
                    fieldnames: list[str] = (
                        list(CRM_COLUMNS)
                        if output_profile == "crm_generico"
                        else list(EXPORT_COLUMNS)
                    )
                    if metadata is not None:
                        fieldnames.extend(self._metadata_columns(metadata))
                    writer = csv.DictWriter(handle, fieldnames=fieldnames)
                    writer.writeheader()
                    for hospital in hospitals:
                        row = self._output_row(hospital, output_profile)
                        row = {
                            key: self._cell_value(value)
                            if isinstance(value, (dict, list, tuple))
                            else value
                            for key, value in row.items()
                        }
                        if metadata is not None:
                            row.update(self._metadata_columns(metadata))
                        writer.writerow(row)
                        records += 1
                    handle.flush()
                    os.fsync(handle.fileno())
            elif normalized_format == "json":
                with temporary.open("w", encoding="utf-8", newline="") as handle:
                    handle.write("[")
                    for hospital in hospitals:
                        if records:
                            handle.write(",")
                        row = (
                            hospital.to_dict()
                            if output_profile is None
                            else self._output_row(hospital, output_profile)
                        )
                        if metadata is not None:
                            row["_metadados"] = dict(metadata)
                        json.dump(
                            row,
                            handle,
                            ensure_ascii=False,
                            separators=(",", ":"),
                        )
                        records += 1
                    handle.write("]")
                    handle.flush()
                    os.fsync(handle.fileno())
            elif normalized_format == "jsonl":
                with temporary.open("w", encoding="utf-8", newline="") as handle:
                    for hospital in hospitals:
                        row = (
                            hospital.to_dict()
                            if output_profile is None
                            else self._output_row(hospital, output_profile)
                        )
                        if metadata is not None:
                            row["_metadados"] = dict(metadata)
                        handle.write(
                            json.dumps(row, ensure_ascii=False, separators=(",", ":")) + "\n"
                        )
                        records += 1
                    handle.flush()
                    os.fsync(handle.fileno())
            else:
                workbook = Workbook(write_only=True)
                sheet = workbook.create_sheet("CNES")
                columns = (
                    list(CRM_COLUMNS) if output_profile == "crm_generico" else list(EXPORT_COLUMNS)
                )
                sheet.append(columns)
                for hospital in hospitals:
                    row = self._output_row(hospital, output_profile)
                    sheet.append(
                        [
                            self._cell_value(row[column])
                            if isinstance(row[column], (dict, list, tuple))
                            else row[column]
                            for column in columns
                        ]
                    )
                    records += 1
                if metadata is not None:
                    metadata_sheet = workbook.create_sheet("_metadados")
                    metadata_sheet.append(["campo", "valor"])
                    for key, value in metadata.items():
                        metadata_sheet.append([key, self._cell_value(value)])
                workbook.save(temporary)
            temporary.replace(output)
        finally:
            if temporary is not None:
                temporary.unlink(missing_ok=True)
        return output, records

    @staticmethod
    def _output_row(hospital: HospitalInfo, output_profile: str | None) -> dict[str, Any]:
        domain_row = hospital.to_dict()
        if output_profile is None:
            return {
                canonical: domain_row[attribute] for canonical, attribute in EXPORT_COLUMNS.items()
            }
        cnpj = domain_row.get("cnpj")
        values = {
            "chave_deduplicacao": (f"{domain_row['cnes']}:{cnpj}" if cnpj else None),
            **{name: domain_row.get(name) for name in CRM_COLUMNS[1:]},
        }
        return {name: values[name] for name in CRM_COLUMNS}

    @staticmethod
    def _metadata_columns(metadata: Mapping[str, Any]) -> dict[str, str]:
        return {
            f"_{key}": LocalDatasetExporter._cell_value(value) for key, value in metadata.items()
        }

    @staticmethod
    def _cell_value(value: Any) -> str:
        if isinstance(value, (dict, list, tuple)):
            return json.dumps(value, ensure_ascii=False, sort_keys=True)
        if value is None:
            return ""
        return str(value)

    @staticmethod
    def _available_output(directory: Path, basename: str, format: str) -> Path:
        candidate = directory / f"{basename}.{format}"
        suffix = 1
        while candidate.exists():
            candidate = directory / f"{basename}-{suffix}.{format}"
            suffix += 1
        return candidate

    def _resolve_destination(self, destination: Path | None) -> Path:
        candidate = self._root if destination is None else destination
        if not candidate.is_absolute():
            candidate = self._root / candidate
        resolved = candidate.resolve(strict=False)
        try:
            resolved.relative_to(self._root)
        except ValueError as exc:
            raise ValueError(
                "destino deve permanecer dentro do diretório de exportação configurado"
            ) from exc
        return resolved
